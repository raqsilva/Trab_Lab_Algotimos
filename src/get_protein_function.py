# -*- coding: utf-8 -*-
# Code to export results from blast to excel table
import os.path
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
#read all txt files from a foulder and return GI protein to a list
GI=[]
for file in os.listdir("../res/blast_without_note/match/allhits/funcao_all_hits/"):
    if file.endswith(".txt"):
        limpo = file.replace(".txt", "")
        GI.append(limpo)

#open file with results of blast and export to excel table
index = 1             
for name in range(len(GI)):
    fic=open("../res/blast_without_note/match/allhits/funcao_all_hits/"+GI[name]+".txt").readlines()
    for i in range(len(fic)) :
        x=fic[i]
        ws['A' + str(index)] = GI[name]
        ws['B' + str(index)] = x[3:10]
        ws['C' + str(index)] = x[29:]
        index+=1
  
wb.save("withoutnote.xlsx")